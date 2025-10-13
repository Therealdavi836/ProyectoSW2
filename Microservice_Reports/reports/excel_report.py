import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_excel(data, filename="reporte.xlsx"):
    """
    Genera un reporte Excel con formato profesional a partir de cualquier JSON.
    - data: lista o dict con datos
    - filename: ruta de guardado
    """
    if isinstance(data, dict):
        data = [data]

    if not data:
        raise ValueError("No se proporcionaron datos para generar el Excel")

    df = pd.DataFrame(data)
    df.to_excel(filename, index=False, engine="openpyxl")

    wb = load_workbook(filename)
    ws = wb.active

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")

    # Estilo para encabezados
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

        max_length = max(
            (len(str(value)) for value in df[column_title].astype(str)),
            default=len(column_title)
        )
        adjusted_width = max(max_length + 2, len(column_title) + 2)
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    # Alinear celdas del cuerpo
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filename)
    return filename